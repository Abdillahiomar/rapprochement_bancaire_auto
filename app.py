"""
====================================================================================
 APPLICATION DE RAPPROCHEMENT DE TRANSACTIONS - D-MONEY x BANQUE PARTENAIRE
====================================================================================
Auteur      : Généré avec Claude (Anthropic) pour Djibouti Telecom / D-Money
Objectif    : Automatiser le rapprochement (réconciliation) des transactions entre
              la plateforme D-Money et le relevé d'une banque partenaire interconnectée.
Technologies: Streamlit, Pandas, Plotly, OpenPyXL

Architecture du code (fonctions principales) :
    - load_data()               -> lecture robuste des CSV (encodage / séparateur)
    - validate_data()            -> contrôles qualité (doublons, ID vides)
    - prepare_dataframe()        -> normalisation des colonnes mappées
    - reconcile_transactions()   -> rapprochement Full Outer Join sur TransactionID
    - compare_amounts()          -> comparaison des montants sur les transactions rapprochées
    - generate_dashboard()       -> affichage des KPI et graphiques Plotly
    - export_to_excel()          -> export multi-feuilles (openpyxl)

Bonnes pratiques appliquées :
    - Fonctions pures séparées de l'affichage Streamlit autant que possible
    - Mise en cache (st.cache_data) des opérations coûteuses et déterministes
    - Gestion des erreurs explicite avec messages clairs pour l'utilisateur
    - Utilisation de st.session_state pour persister les résultats entre les reruns
    - Typage indicatif et docstrings sur toutes les fonctions
====================================================================================
"""

import io
import csv
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ====================================================================================
# CONFIGURATION GENERALE DE LA PAGE
# ====================================================================================

st.set_page_config(
    page_title="Rapprochement D-Money x Banque",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Palette de marque D-Money (navy / gold) pour une identité visuelle cohérente
COLOR_NAVY = "#1B2F6E"
COLOR_GOLD = "#F5A800"
COLOR_GREEN = "#1E8E5A"
COLOR_RED = "#C0392B"
COLOR_GRAY = "#6C757D"

CUSTOM_CSS = f"""
<style>
    .main {{
        background-color: #F7F8FA;
    }}
    .kpi-card {{
        background-color: white;
        border-radius: 10px;
        padding: 1.1rem 1.2rem;
        border-left: 5px solid {COLOR_NAVY};
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    }}
    .kpi-card h3 {{
        margin: 0;
        font-size: 0.85rem;
        color: {COLOR_GRAY};
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.03em;
    }}
    .kpi-card p {{
        margin: 0.2rem 0 0 0;
        font-size: 1.7rem;
        font-weight: 700;
        color: {COLOR_NAVY};
    }}
    div.stButton > button:first-child {{
        background-color: {COLOR_NAVY};
        color: white;
        font-weight: 600;
        border-radius: 6px;
        border: none;
    }}
    div.stButton > button:first-child:hover {{
        background-color: {COLOR_GOLD};
        color: {COLOR_NAVY};
    }}
    section[data-testid="stSidebar"] {{
        background-color: {COLOR_NAVY};
    }}
    section[data-testid="stSidebar"] * {{
        color: white !important;
    }}
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ====================================================================================
# 1. CHARGEMENT DES DONNEES (encodage UTF-8 / ISO-8859-1, séparateur auto-détecté)
# ====================================================================================

def _detect_separator(sample_text: str) -> str:
    """Détecte le séparateur CSV (',', ';', '\\t', '|') à partir d'un échantillon de texte."""
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=[",", ";", "\t", "|"])
        return dialect.delimiter
    except csv.Error:
        # Repli raisonnable : le point-virgule est très répandu dans les exports bancaires FR
        if ";" in sample_text.splitlines()[0] if sample_text else False:
            return ";"
        return ","


def _find_header_row(full_text: str, sep: str, max_scan: int = 15) -> int:
    """
    Détecte l'index de la ligne d'en-tête réelle dans un export contenant des lignes
    de métadonnées avant le tableau de données (très fréquent dans les relevés
    bancaires et les exports D-Money : "Account Holder:", "Opening Balance:",
    "Time Period:", etc.).

    Principe : on essaie de lire le fichier avec un nombre croissant de lignes
    ignorées (skiprows = 0, 1, 2, ...). Tant que la ligne d'en-tête utilisée ne
    correspond pas à la vraie ligne de colonnes, soit la lecture échoue (nombre
    de champs incohérent avec les lignes suivantes), soit pandas génère des
    colonnes "Unnamed: N" (cellules d'en-tête vides). On retient le premier
    skiprows qui produit un DataFrame propre : au moins 2 colonnes, aucune
    colonne "Unnamed".

    Args:
        full_text: contenu texte complet du fichier (déjà décodé).
        sep: séparateur CSV détecté.
        max_scan: nombre maximal de lignes de préambule testées.

    Returns:
        Index (0-based) de la ligne d'en-tête à utiliser avec `skiprows` dans pandas.
    """
    for skip in range(max_scan + 1):
        try:
            probe = pd.read_csv(
                io.StringIO(full_text),
                sep=sep,
                engine="c",
                skiprows=skip,
                nrows=3,
                dtype=str,
                keep_default_na=False,
            )
        except (pd.errors.ParserError, UnicodeDecodeError):
            continue

        cols = list(probe.columns)
        if len(cols) < 2:
            continue
        if any(str(c).strip() == "" or str(c).startswith("Unnamed") for c in cols):
            continue
        if not isinstance(probe.index, pd.RangeIndex):
            # Les lignes suivantes ont plus de champs que l'en-tête testé : pandas a
            # replié les colonnes excédentaires dans un (Multi)Index -> mauvais en-tête.
            continue
        return skip

    return 0


@st.cache_data(show_spinner=False)
def load_data(file_bytes: bytes, file_name: str):
    """
    Charge un fichier CSV en gérant automatiquement l'encodage, le séparateur,
    et les éventuelles lignes de métadonnées précédant le tableau de données.

    Essaie successivement les encodages UTF-8 (avec/sans BOM) puis ISO-8859-1 / CP1252,
    qui couvrent la quasi-totalité des exports bancaires et des systèmes D-Money.

    Args:
        file_bytes: contenu brut du fichier (bytes).
        file_name: nom du fichier (utilisé uniquement pour les messages d'erreur).

    Returns:
        tuple (DataFrame, encodage_utilisé, séparateur_utilisé, nb_lignes_ignorées)

    Raises:
        ValueError: si le fichier ne peut être lu avec aucun des encodages testés.
    """
    encodings_to_try = ["utf-8-sig", "utf-8", "iso-8859-1", "cp1252"]
    last_error = None

    for encoding in encodings_to_try:
        try:
            full_text = file_bytes.decode(encoding, errors="strict")
            sample = full_text[:8192]
            separator = _detect_separator(sample)
            header_row = _find_header_row(full_text, separator)

            df = pd.read_csv(
                io.StringIO(full_text),
                sep=separator,
                engine="c",
                dtype=str,          # on lit tout en texte, les conversions sont faites plus tard
                keep_default_na=False,
                na_values=["", "NA", "N/A", "null", "NULL", "None"],
                skiprows=header_row,
            )
            if df.shape[1] <= 1:
                raise ValueError("Aucune colonne exploitable détectée (vérifiez le séparateur).")
            # Colonnes potentiellement dupliquées ou totalement vides (ex: colonnes "Currency"
            # répétées dans certains relevés bancaires) sont conservées telles quelles ;
            # pandas les distingue automatiquement (Currency, Currency.1, ...).
            return df, encoding, separator, header_row
        except (UnicodeDecodeError, pd.errors.ParserError, ValueError) as exc:
            last_error = exc
            continue

    raise ValueError(
        f"Impossible de lire le fichier '{file_name}'. "
        f"Vérifiez qu'il s'agit bien d'un fichier CSV valide (UTF-8 ou ISO-8859-1). "
        f"Erreur technique : {last_error}"
    )


# ====================================================================================
# 2. CONTROLES QUALITE
# ====================================================================================

def validate_data(df: pd.DataFrame, id_col: str) -> dict:
    """
    Effectue les contrôles qualité sur un fichier de transactions.

    Vérifie les identifiants (TransactionID) vides et les doublons.

    Args:
        df: DataFrame source.
        id_col: nom de la colonne identifiant à contrôler.

    Returns:
        Dictionnaire de statistiques et de sous-DataFrames d'anomalies.
    """
    total = len(df)
    id_series = df[id_col].astype(str).str.strip()

    null_mask = df[id_col].isna() | (id_series == "") | (id_series.str.lower() == "nan")
    null_count = int(null_mask.sum())

    non_null = ~null_mask
    dup_mask = df[id_col].duplicated(keep=False) & non_null
    dup_count = int(dup_mask.sum())
    dup_unique_ids = int(df.loc[dup_mask, id_col].nunique())

    return {
        "total": total,
        "null_count": null_count,
        "valid_count": total - null_count,
        "duplicate_row_count": dup_count,
        "duplicate_unique_ids": dup_unique_ids,
        "duplicate_df": df.loc[dup_mask].sort_values(id_col),
        "null_df": df.loc[null_mask],
        "quality_rate": round((total - null_count - dup_count) / total * 100, 2) if total else 0.0,
    }


def extract_id_from_text(series: pd.Series, separator: str, position: str = "last") -> pd.Series:
    """
    Extrait un identifiant depuis une colonne texte libre, à partir d'un séparateur.

    Utile lorsque le TransactionID n'est pas isolé dans une colonne dédiée mais
    intégré dans un champ descriptif, par exemple :
    "Deposit From D-money - 000370600685" -> "000370600685".

    Args:
        series: colonne source (texte libre).
        separator: caractère ou chaîne séparateur (ex: "-").
        position: "last" pour prendre la partie après le DERNIER séparateur
                  (recommandé dans la majorité des cas),
                  "first" pour la partie après le PREMIER séparateur.

    Returns:
        Series de chaînes extraites et nettoyées. Renvoie None (valeur manquante)
        quand le séparateur est absent du texte (ex: "DMONEY TRANSFER"), ce qui
        est ensuite correctement détecté comme TransactionID vide par validate_data().
    """
    def _extract(value):
        if pd.isna(value):
            return None
        text = str(value)
        if separator not in text:
            return None
        part = text.rsplit(separator, 1)[-1] if position == "last" else text.split(separator, 1)[-1]
        return part.strip()

    return series.apply(_extract)


# ====================================================================================
# 3. MAPPING / NORMALISATION DES COLONNES
# ====================================================================================

def prepare_dataframe(
    df: pd.DataFrame,
    id_col: str,
    amount_col: str | None,
    date_col: str | None,
    status_col: str | None,
    ref_col: str | None,
    side: str,
) -> pd.DataFrame:
    """
    Normalise un DataFrame selon le mapping de colonnes choisi par l'utilisateur.

    Renomme les colonnes sélectionnées vers des noms standards suffixés par `side`
    ("DM" ou "BANK"), afin d'éviter toute ambiguïté lors du merge, y compris
    lorsque les deux fichiers utilisent des noms de colonnes identiques.

    Args:
        df: DataFrame source (D-Money ou Banque).
        id_col: nom de la colonne TransactionID dans df.
        amount_col, date_col, status_col, ref_col: colonnes optionnelles (ou None).
        side: "DM" ou "BANK".

    Returns:
        DataFrame normalisé avec colonnes standardisées + 'TransactionID'.
    """
    work = df.copy()
    work["TransactionID"] = work[id_col].astype(str).str.strip()

    rename_map = {}
    if amount_col:
        rename_map[amount_col] = f"Amount_{side}"
    if date_col:
        rename_map[date_col] = f"Date_{side}"
    if status_col:
        rename_map[status_col] = f"Status_{side}"
    if ref_col:
        rename_map[ref_col] = f"ExtRef_{side}"

    work = work.rename(columns=rename_map)

    keep_cols = ["TransactionID"] + [c for c in rename_map.values() if c in work.columns]
    # On conserve aussi toutes les colonnes originales (préfixées) pour l'export détaillé
    other_cols = [c for c in df.columns if c not in [id_col, amount_col, date_col, status_col, ref_col]]
    for c in other_cols:
        new_name = f"{c}_{side}_orig"
        work[new_name] = df[c]
        keep_cols.append(new_name)

    return work[keep_cols]


# ====================================================================================
# 4. RAPPROCHEMENT (FULL OUTER JOIN)
# ====================================================================================

def reconcile_transactions(dm_ready: pd.DataFrame, bank_ready: pd.DataFrame) -> dict:
    """
    Effectue le rapprochement Full Outer Join sur la colonne 'TransactionID'.

    Les identifiants vides sont exclus en amont (cf validate_data). Pour les
    doublons de TransactionID, seule la première occurrence est utilisée pour
    le rapprochement (les doublons sont déjà signalés séparément en contrôle qualité)
    afin d'éviter un produit cartésien qui fausserait les statistiques.

    Args:
        dm_ready: DataFrame D-Money normalisé (sortie de prepare_dataframe).
        bank_ready: DataFrame Banque normalisé (sortie de prepare_dataframe).

    Returns:
        dict contenant : matched, only_dm, only_bank, merged (DataFrame complet)
    """
    dm_valid = dm_ready[dm_ready["TransactionID"].str.len() > 0].drop_duplicates(
        subset="TransactionID", keep="first"
    )
    bank_valid = bank_ready[bank_ready["TransactionID"].str.len() > 0].drop_duplicates(
        subset="TransactionID", keep="first"
    )

    merged = dm_valid.merge(
        bank_valid, on="TransactionID", how="outer", indicator=True
    )

    matched = merged[merged["_merge"] == "both"].drop(columns="_merge").copy()
    only_dm = merged[merged["_merge"] == "left_only"].drop(columns="_merge").copy()
    only_bank = merged[merged["_merge"] == "right_only"].drop(columns="_merge").copy()

    # Nettoyage : on retire les colonnes provenant de l'autre fichier qui sont vides
    # dans le sous-ensemble "only" (purement esthétique pour l'affichage/export)
    dm_only_cols = [c for c in only_dm.columns if c.endswith("_BANK") or c.endswith("_BANK_orig")]
    only_dm = only_dm.drop(columns=dm_only_cols, errors="ignore")

    bank_only_cols = [c for c in only_bank.columns if c.endswith("_DM") or c.endswith("_DM_orig")]
    only_bank = only_bank.drop(columns=bank_only_cols, errors="ignore")

    return {
        "matched": matched,
        "only_dm": only_dm,
        "only_bank": only_bank,
        "merged_full": merged.drop(columns="_merge"),
    }


# ====================================================================================
# 5. ANALYSE DES MONTANTS
# ====================================================================================

def compare_amounts(matched_df: pd.DataFrame, tolerance: float = 0.01) -> pd.DataFrame | None:
    """
    Compare les montants D-Money et Banque pour les transactions rapprochées.

    Args:
        matched_df: DataFrame des transactions rapprochées (issu de reconcile_transactions).
        tolerance: écart maximal (en valeur absolue) toléré pour considérer un montant
                   comme identique (gère les problèmes d'arrondi).

    Returns:
        DataFrame enrichi avec les colonnes 'Ecart_Montant' et 'Statut_Montant',
        ou None si les colonnes de montant n'ont pas été mappées.
    """
    if "Amount_DM" not in matched_df.columns or "Amount_BANK" not in matched_df.columns:
        return None

    df = matched_df.copy()
    df["Amount_DM_num"] = pd.to_numeric(
        df["Amount_DM"].astype(str).str.replace(" ", "").str.replace(",", "."), errors="coerce"
    )
    df["Amount_BANK_num"] = pd.to_numeric(
        df["Amount_BANK"].astype(str).str.replace(" ", "").str.replace(",", "."), errors="coerce"
    )
    df["Ecart_Montant"] = df["Amount_DM_num"] - df["Amount_BANK_num"]

    df["Statut_Montant"] = np.where(
        df["Ecart_Montant"].abs() <= tolerance, "Identique", "Différent"
    )
    non_comparable = df["Amount_DM_num"].isna() | df["Amount_BANK_num"].isna()
    df.loc[non_comparable, "Statut_Montant"] = "Non comparable"

    return df


# ====================================================================================
# 6. EXPORT EXCEL (OPENPYXL)
# ====================================================================================

def _style_worksheet(ws):
    """Applique une mise en forme professionnelle à une feuille Excel (en-tête, largeur, filtre)."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2F6E", end_color="1B2F6E", fill_type="solid")

    if ws.max_row == 0:
        return

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(ws.max_row, 500) + 1):  # échantillon pour la perf
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def export_to_excel(sheets: dict, kpis: dict) -> io.BytesIO:
    """
    Génère un classeur Excel multi-feuilles avec mise en forme professionnelle.

    Args:
        sheets: dict {nom_de_feuille: DataFrame} pour les feuilles de détail.
        kpis: dict {libellé: valeur} pour la feuille "Résumé exécutif".

    Returns:
        Buffer BytesIO prêt à être proposé au téléchargement.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Feuille de résumé exécutif en première position
        resume_df = pd.DataFrame(list(kpis.items()), columns=["Indicateur", "Valeur"])
        resume_df.to_excel(writer, sheet_name="Résumé exécutif", index=False)

        for name, df in sheets.items():
            safe_name = name[:31]  # limite Excel = 31 caractères
            (df if not df.empty else pd.DataFrame({"Info": ["Aucune donnée"]})).to_excel(
                writer, sheet_name=safe_name, index=False
            )

        for ws in writer.sheets.values():
            _style_worksheet(ws)

    output.seek(0)
    return output


# ====================================================================================
# 7. DASHBOARD / VISUALISATIONS
# ====================================================================================

def kpi_card(col, label: str, value, color=COLOR_NAVY):
    """Affiche une carte KPI stylisée dans la colonne Streamlit donnée."""
    col.markdown(
        f"""
        <div class="kpi-card" style="border-left-color:{color};">
            <h3>{label}</h3>
            <p style="color:{color};">{value}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def generate_dashboard(stats: dict, results: dict, amounts_df: pd.DataFrame | None):
    """
    Affiche le tableau de bord de synthèse : KPI + graphiques Plotly.

    Args:
        stats: dict des compteurs globaux (total_dm, total_bank, matched, ecarts, taux).
        results: dict issu de reconcile_transactions (pour les graphiques de répartition).
        amounts_df: DataFrame issu de compare_amounts (ou None si non applicable).
    """
    st.subheader("📊 Indicateurs clés de performance")
    c1, c2, c3, c4, c5 = st.columns(5)
    kpi_card(c1, "Transactions D-Money", f"{stats['total_dm']:,}".replace(",", " "))
    kpi_card(c2, "Transactions Banque", f"{stats['total_bank']:,}".replace(",", " "))
    kpi_card(c3, "Rapprochées", f"{stats['matched']:,}".replace(",", " "), COLOR_GREEN)
    kpi_card(c4, "Écarts", f"{stats['ecarts']:,}".replace(",", " "), COLOR_RED)
    kpi_card(c5, "Taux de rapprochement", f"{stats['taux']} %", COLOR_GOLD)

    st.divider()

    col_a, col_b = st.columns(2)

    with col_a:
        pie_df = pd.DataFrame(
            {
                "Catégorie": ["Rapprochées", "Uniquement D-Money", "Uniquement Banque"],
                "Nombre": [
                    len(results["matched"]),
                    len(results["only_dm"]),
                    len(results["only_bank"]),
                ],
            }
        )
        fig_pie = px.pie(
            pie_df,
            names="Catégorie",
            values="Nombre",
            title="Répartition des transactions",
            color="Catégorie",
            color_discrete_map={
                "Rapprochées": COLOR_GREEN,
                "Uniquement D-Money": COLOR_GOLD,
                "Uniquement Banque": COLOR_NAVY,
            },
            hole=0.45,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_b:
        bar_df = pd.DataFrame(
            {
                "Type d'écart": ["Uniquement D-Money", "Uniquement Banque"],
                "Nombre": [len(results["only_dm"]), len(results["only_bank"])],
            }
        )
        fig_bar = px.bar(
            bar_df,
            x="Type d'écart",
            y="Nombre",
            title="Écarts détectés par type",
            color="Type d'écart",
            color_discrete_map={
                "Uniquement D-Money": COLOR_GOLD,
                "Uniquement Banque": COLOR_NAVY,
            },
            text="Nombre",
        )
        fig_bar.update_traces(textposition="outside")
        fig_bar.update_layout(showlegend=False)
        st.plotly_chart(fig_bar, use_container_width=True)

    if amounts_df is not None and not amounts_df.empty:
        st.divider()
        st.subheader("💰 Analyse des montants")

        total_dm_amount = amounts_df["Amount_DM_num"].sum()
        total_bank_amount = amounts_df["Amount_BANK_num"].sum()
        ecart_total = total_dm_amount - total_bank_amount
        incoherent_count = int((amounts_df["Statut_Montant"] == "Différent").sum())

        c1, c2, c3, c4 = st.columns(4)
        kpi_card(c1, "Montant total D-Money", f"{total_dm_amount:,.2f}".replace(",", " "))
        kpi_card(c2, "Montant total Banque", f"{total_bank_amount:,.2f}".replace(",", " "))
        kpi_card(c3, "Écart total", f"{ecart_total:,.2f}".replace(",", " "), COLOR_RED if abs(ecart_total) > 0.01 else COLOR_GREEN)
        kpi_card(c4, "Montants incohérents", f"{incoherent_count:,}".replace(",", " "), COLOR_RED)

        col_c, col_d = st.columns(2)
        with col_c:
            status_counts = amounts_df["Statut_Montant"].value_counts().reset_index()
            status_counts.columns = ["Statut", "Nombre"]
            fig_status = px.pie(
                status_counts,
                names="Statut",
                values="Nombre",
                title="Cohérence des montants",
                color="Statut",
                color_discrete_map={
                    "Identique": COLOR_GREEN,
                    "Différent": COLOR_RED,
                    "Non comparable": COLOR_GRAY,
                },
                hole=0.45,
            )
            st.plotly_chart(fig_status, use_container_width=True)

        with col_d:
            diffs = amounts_df.loc[amounts_df["Statut_Montant"] == "Différent", "Ecart_Montant"]
            if not diffs.empty:
                fig_hist = px.histogram(
                    diffs,
                    nbins=30,
                    title="Distribution des écarts de montant",
                    color_discrete_sequence=[COLOR_NAVY],
                )
                fig_hist.update_layout(showlegend=False, xaxis_title="Écart (D-Money - Banque)", yaxis_title="Fréquence")
                st.plotly_chart(fig_hist, use_container_width=True)
            else:
                st.info("Aucun écart de montant à afficher : tous les montants rapprochés sont identiques.")


# ====================================================================================
# INTERFACE UTILISATEUR STREAMLIT
# ====================================================================================

def main():
    st.title("🏦 Rapprochement de transactions — D-Money x Banque partenaire")
    st.caption(
        "Automatisez le contrôle et le rapprochement des transactions interconnectées "
        "entre D-Money et le relevé bancaire du partenaire."
    )

    if "results" not in st.session_state:
        st.session_state.results = None

    # --------------------------------------------------------------------------
    # SIDEBAR : upload + réglages
    # --------------------------------------------------------------------------
    with st.sidebar:
        st.header("⚙️ Paramètres")
        tolerance = st.number_input(
            "Tolérance d'écart de montant (FDJ)",
            min_value=0.0, max_value=100.0, value=0.01, step=0.01,
            help="Écart maximal toléré entre les montants D-Money et Banque pour être considéré comme identique.",
        )
        st.divider()
        if st.button("🔄 Réinitialiser l'application"):
            st.session_state.results = None
            st.rerun()

    # --------------------------------------------------------------------------
    # ETAPE 1 : IMPORT DES FICHIERS
    # --------------------------------------------------------------------------
    st.header("1️⃣ Import des fichiers")
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        dm_file = st.file_uploader("Fichier D-Money (CSV)", type=["csv"], key="dm_upload")
    with col_up2:
        bank_file = st.file_uploader("Fichier Banque partenaire (CSV)", type=["csv"], key="bank_upload")

    if not dm_file or not bank_file:
        st.info("⬆️ Veuillez importer les deux fichiers CSV (D-Money et Banque) pour continuer.")
        st.subheader("📄 Structure de fichier attendue (exemple)")
        st.markdown("**Fichier D-Money** — colonnes typiques : `TransactionID`, `Date`, `Montant`, `Statut`, `Referenceexterne`")
        st.code(
            "TransactionID,Date,Montant,Statut,ReferenceExterne\n"
            "DM100001,2026-06-01,15000,SUCCESS,BK-2201\n"
            "DM100002,2026-06-01,7500,SUCCESS,BK-2202\n"
            "DM100003,2026-06-02,32000,FAILED,\n",
            language="csv",
        )
        st.markdown("**Fichier Banque** — les noms de colonnes peuvent différer, d'où le mapping dynamique ci-dessous :")
        st.code(
            "IdTransactionDMoney;DateOperation;MontantOperation;EtatOperation\n"
            "DM100001;01/06/2026;15000.00;VALIDEE\n"
            "DM100002;01/06/2026;7500.00;VALIDEE\n"
            "DM100004;02/06/2026;9800.00;VALIDEE\n",
            language="csv",
        )
        return

    # --------------------------------------------------------------------------
    # CHARGEMENT ROBUSTE DES FICHIERS
    # --------------------------------------------------------------------------
    try:
        df_dm, enc_dm, sep_dm, skipped_dm = load_data(dm_file.getvalue(), dm_file.name)
        df_bank, enc_bank, sep_bank, skipped_bank = load_data(bank_file.getvalue(), bank_file.name)
    except ValueError as exc:
        st.error(f"❌ Erreur de lecture : {exc}")
        return

    col_prev1, col_prev2 = st.columns(2)
    with col_prev1:
        skip_msg = f" | {skipped_dm} ligne(s) d'en-tête ignorée(s)" if skipped_dm else ""
        st.success(f"✅ D-Money chargé — {len(df_dm):,} lignes | encodage `{enc_dm}` | séparateur `{sep_dm}`{skip_msg}".replace(",", " "))
        st.dataframe(df_dm.head(10), use_container_width=True, height=200)
    with col_prev2:
        skip_msg = f" | {skipped_bank} ligne(s) d'en-tête ignorée(s)" if skipped_bank else ""
        st.success(f"✅ Banque chargée — {len(df_bank):,} lignes | encodage `{enc_bank}` | séparateur `{sep_bank}`{skip_msg}".replace(",", " "))
        st.dataframe(df_bank.head(10), use_container_width=True, height=200)

    if df_dm.empty or df_bank.empty:
        st.error("❌ Un des deux fichiers ne contient aucune ligne exploitable.")
        return

    # --------------------------------------------------------------------------
    # ETAPE 2 : MAPPING DES COLONNES
    # --------------------------------------------------------------------------
    st.header("2️⃣ Mapping des colonnes")
    st.caption("Sélectionnez la correspondance des colonnes entre les deux fichiers. Seul le TransactionID est obligatoire.")

    NONE_OPTION = "— Aucune —"
    MODE_DIRECT = "Colonne directe"
    MODE_EXTRACT = "Extraire depuis un texte (ex: après un séparateur)"

    def _id_mapping_widgets(df: pd.DataFrame, side_key: str, side_label: str):
        """
        Affiche les widgets de mapping du TransactionID pour un fichier donné,
        avec deux modes possibles :
          - Colonne directe : l'identifiant est déjà isolé dans une colonne.
          - Extraction : l'identifiant est noyé dans un champ texte libre
            (ex: "Deposit From D-money - 000370600685"), et doit être extrait
            via un séparateur.

        Returns:
            tuple (df_travail, nom_colonne_id_a_utiliser)
        """
        mode = st.radio(
            f"Méthode d'identification — {side_label}",
            options=[MODE_DIRECT, MODE_EXTRACT],
            key=f"{side_key}_id_mode",
            horizontal=True,
        )

        if mode == MODE_DIRECT:
            id_col = st.selectbox("Colonne TransactionID (obligatoire)", options=df.columns, key=f"{side_key}_id_direct")
            return df, id_col

        source_col = st.selectbox("Colonne source contenant l'identifiant", options=df.columns, key=f"{side_key}_id_source")
        col_sep, col_pos = st.columns(2)
        with col_sep:
            separator = st.text_input("Séparateur", value="-", key=f"{side_key}_id_sep")
        with col_pos:
            position_label = st.selectbox(
                "Partie à extraire",
                options=["Après le dernier séparateur", "Après le premier séparateur"],
                key=f"{side_key}_id_pos",
            )
        position = "last" if position_label.startswith("Après le dernier") else "first"

        work_df = df.copy()
        extracted_col = "_TransactionID_extracted"
        if separator:
            work_df[extracted_col] = extract_id_from_text(work_df[source_col], separator, position)
            preview = work_df[[source_col, extracted_col]].head(4)
            st.caption("Aperçu de l'extraction :")
            st.dataframe(preview, use_container_width=True, height=150)
            missing = work_df[extracted_col].isna().sum()
            if missing:
                st.caption(f"ℹ️ {missing} ligne(s) sans séparateur trouvé — seront comptées comme ID vide.")
        else:
            st.warning("⚠️ Veuillez renseigner un séparateur.")
            work_df[extracted_col] = None

        return work_df, extracted_col

    col_map1, col_map2 = st.columns(2)

    with col_map1:
        st.markdown("**Fichier D-Money**")
        df_dm, dm_id_col = _id_mapping_widgets(df_dm, "dm", "D-Money")
        dm_amount_col = st.selectbox("Montant", options=[NONE_OPTION] + list(df_dm.columns), key="dm_amount")
        dm_date_col = st.selectbox("Date transaction", options=[NONE_OPTION] + list(df_dm.columns), key="dm_date")
        dm_status_col = st.selectbox("Statut", options=[NONE_OPTION] + list(df_dm.columns), key="dm_status")
        dm_ref_col = st.selectbox("Référence externe", options=[NONE_OPTION] + list(df_dm.columns), key="dm_ref")

    with col_map2:
        st.markdown("**Fichier Banque**")
        df_bank, bank_id_col = _id_mapping_widgets(df_bank, "bank", "Banque")
        bank_amount_col = st.selectbox("Montant", options=[NONE_OPTION] + list(df_bank.columns), key="bank_amount")
        bank_date_col = st.selectbox("Date transaction", options=[NONE_OPTION] + list(df_bank.columns), key="bank_date")
        bank_status_col = st.selectbox("Statut", options=[NONE_OPTION] + list(df_bank.columns), key="bank_status")
        bank_ref_col = st.selectbox("Référence externe", options=[NONE_OPTION] + list(df_bank.columns), key="bank_ref")

    def _clean(v):
        return None if v == NONE_OPTION else v

    dm_amount_col, dm_date_col, dm_status_col, dm_ref_col = map(_clean, [dm_amount_col, dm_date_col, dm_status_col, dm_ref_col])
    bank_amount_col, bank_date_col, bank_status_col, bank_ref_col = map(_clean, [bank_amount_col, bank_date_col, bank_status_col, bank_ref_col])

    # --------------------------------------------------------------------------
    # ETAPE 3 : LANCEMENT DU RAPPROCHEMENT
    # --------------------------------------------------------------------------
    st.header("3️⃣ Lancement du rapprochement")

    if st.button("🚀 Lancer le rapprochement des transactions", type="primary"):
        progress = st.progress(0, text="Initialisation...")

        try:
            progress.progress(10, text="Contrôle qualité des données...")
            validation_dm = validate_data(df_dm, dm_id_col)
            validation_bank = validate_data(df_bank, bank_id_col)

            progress.progress(35, text="Normalisation des colonnes mappées...")
            dm_ready = prepare_dataframe(df_dm, dm_id_col, dm_amount_col, dm_date_col, dm_status_col, dm_ref_col, side="DM")
            bank_ready = prepare_dataframe(df_bank, bank_id_col, bank_amount_col, bank_date_col, bank_status_col, bank_ref_col, side="BANK")

            progress.progress(60, text="Rapprochement Full Outer Join en cours...")
            results = reconcile_transactions(dm_ready, bank_ready)

            progress.progress(80, text="Analyse des montants...")
            amounts_df = compare_amounts(results["matched"], tolerance=tolerance)

            progress.progress(95, text="Finalisation...")

            matched_count = len(results["matched"])
            only_dm_count = len(results["only_dm"])
            only_bank_count = len(results["only_bank"])
            union_count = matched_count + only_dm_count + only_bank_count
            taux = round(matched_count / union_count * 100, 2) if union_count else 0.0

            stats = {
                "total_dm": len(df_dm),
                "total_bank": len(df_bank),
                "matched": matched_count,
                "ecarts": only_dm_count + only_bank_count,
                "taux": taux,
            }

            st.session_state.results = {
                "results": results,
                "amounts_df": amounts_df,
                "stats": stats,
                "validation_dm": validation_dm,
                "validation_bank": validation_bank,
                "has_amount": dm_amount_col is not None and bank_amount_col is not None,
                "has_date": dm_date_col is not None and bank_date_col is not None,
            }

            progress.progress(100, text="Terminé !")
            st.success("✅ Rapprochement terminé avec succès.")
        except KeyError as exc:
            st.error(f"❌ Colonne manquante lors du traitement : {exc}")
            return
        except Exception as exc:  # garde-fou générique avec message clair
            st.error(f"❌ Une erreur inattendue est survenue durant le rapprochement : {exc}")
            return

    # --------------------------------------------------------------------------
    # AFFICHAGE DES RESULTATS (si déjà calculés)
    # --------------------------------------------------------------------------
    if st.session_state.results is None:
        return

    state = st.session_state.results
    results = state["results"]
    amounts_df = state["amounts_df"]
    stats = state["stats"]
    validation_dm = state["validation_dm"]
    validation_bank = state["validation_bank"]

    st.divider()

    tabs = st.tabs([
        "📊 Dashboard",
        "✅ Rapprochées",
        "🟠 Uniquement D-Money",
        "🔵 Uniquement Banque",
        "⚠️ Qualité & Doublons",
        "💰 Analyse des montants",
        "🔍 Recherche avancée",
        "📥 Export Excel",
    ])

    # --- Onglet Dashboard --------------------------------------------------
    with tabs[0]:
        generate_dashboard(stats, results, amounts_df)

    # --- Onglet Rapprochées --------------------------------------------------
    with tabs[1]:
        st.subheader(f"✅ Transactions rapprochées ({len(results['matched']):,})".replace(",", " "))
        st.dataframe(results["matched"], use_container_width=True)

    # --- Onglet Uniquement D-Money ------------------------------------------
    with tabs[2]:
        st.subheader(f"🟠 Uniquement dans D-Money ({len(results['only_dm']):,})".replace(",", " "))
        st.caption("Transactions présentes chez D-Money mais absentes du relevé bancaire — à investiguer.")
        st.dataframe(results["only_dm"], use_container_width=True)

    # --- Onglet Uniquement Banque --------------------------------------------
    with tabs[3]:
        st.subheader(f"🔵 Uniquement dans la Banque ({len(results['only_bank']):,})".replace(",", " "))
        st.caption("Transactions présentes dans le relevé bancaire mais absentes de D-Money — à investiguer.")
        st.dataframe(results["only_bank"], use_container_width=True)

    # --- Onglet Qualité & Doublons -------------------------------------------
    with tabs[4]:
        st.subheader("⚠️ Qualité des données")
        col_q1, col_q2 = st.columns(2)
        with col_q1:
            st.markdown("**Fichier D-Money**")
            kpi_card(st, "Taux de qualité", f"{validation_dm['quality_rate']} %")
            st.write(f"- ID vides : **{validation_dm['null_count']}**")
            st.write(f"- Lignes en doublon : **{validation_dm['duplicate_row_count']}** ({validation_dm['duplicate_unique_ids']} ID distincts)")
        with col_q2:
            st.markdown("**Fichier Banque**")
            kpi_card(st, "Taux de qualité", f"{validation_bank['quality_rate']} %")
            st.write(f"- ID vides : **{validation_bank['null_count']}**")
            st.write(f"- Lignes en doublon : **{validation_bank['duplicate_row_count']}** ({validation_bank['duplicate_unique_ids']} ID distincts)")

        st.divider()
        st.markdown("**Détail des doublons — D-Money**")
        st.dataframe(validation_dm["duplicate_df"], use_container_width=True)
        st.markdown("**Détail des doublons — Banque**")
        st.dataframe(validation_bank["duplicate_df"], use_container_width=True)
        st.markdown("**TransactionID vides — D-Money**")
        st.dataframe(validation_dm["null_df"], use_container_width=True)
        st.markdown("**TransactionID vides — Banque**")
        st.dataframe(validation_bank["null_df"], use_container_width=True)

    # --- Onglet Analyse des montants ------------------------------------------
    with tabs[5]:
        if amounts_df is None:
            st.info("ℹ️ Mappez une colonne Montant pour D-Money et Banque pour activer cette analyse.")
        else:
            st.subheader("💰 Détail des écarts de montant")
            statut_filter = st.multiselect(
                "Filtrer par statut",
                options=["Identique", "Différent", "Non comparable"],
                default=["Différent", "Non comparable"],
            )
            filtered = amounts_df[amounts_df["Statut_Montant"].isin(statut_filter)] if statut_filter else amounts_df
            st.dataframe(filtered, use_container_width=True)

    # --- Onglet Recherche avancée ---------------------------------------------
    with tabs[6]:
        st.subheader("🔍 Recherche avancée")
        source_choice = st.selectbox(
            "Rechercher dans",
            options=["Toutes les transactions (fusion complète)", "Rapprochées", "Uniquement D-Money", "Uniquement Banque"],
        )
        source_map = {
            "Toutes les transactions (fusion complète)": results["merged_full"],
            "Rapprochées": results["matched"],
            "Uniquement D-Money": results["only_dm"],
            "Uniquement Banque": results["only_bank"],
        }
        base_df = source_map[source_choice]

        col_s1, col_s2 = st.columns(2)
        with col_s1:
            search_id = st.text_input("Recherche par TransactionID (contient)")
        with col_s2:
            amount_cols_present = [c for c in ["Amount_DM", "Amount_BANK"] if c in base_df.columns]
            search_amount = None
            if amount_cols_present:
                search_amount = st.text_input("Recherche par montant exact (optionnel)")

        filtered_df = base_df.copy()
        if search_id:
            filtered_df = filtered_df[filtered_df["TransactionID"].str.contains(search_id, case=False, na=False)]
        if search_amount:
            try:
                target = float(search_amount.replace(",", "."))
                mask = pd.Series(False, index=filtered_df.index)
                for c in amount_cols_present:
                    mask |= pd.to_numeric(filtered_df[c], errors="coerce").round(2) == round(target, 2)
                filtered_df = filtered_df[mask]
            except ValueError:
                st.warning("⚠️ Montant recherché invalide — utilisez un format numérique (ex: 15000.50).")

        date_cols_present = [c for c in ["Date_DM", "Date_BANK"] if c in base_df.columns]
        if date_cols_present:
            date_text = st.text_input("Recherche par date (contient, ex: 2026-06-01)")
            if date_text:
                mask = pd.Series(False, index=filtered_df.index)
                for c in date_cols_present:
                    mask |= filtered_df[c].astype(str).str.contains(date_text, case=False, na=False)
                filtered_df = filtered_df[mask]

        st.write(f"**{len(filtered_df):,}** résultat(s) trouvé(s).".replace(",", " "))
        st.dataframe(filtered_df, use_container_width=True)

    # --- Onglet Export Excel ---------------------------------------------------
    with tabs[7]:
        st.subheader("📥 Export des résultats (Excel)")
        st.caption("Le fichier généré contient 6 feuilles : Résumé exécutif, Rapprochées, Uniquement D-Money, "
                    "Uniquement Banque, Différences de montants, Doublons.")

        diff_amounts_sheet = (
            amounts_df[amounts_df["Statut_Montant"] == "Différent"]
            if amounts_df is not None
            else pd.DataFrame({"Info": ["Aucune colonne de montant mappée"]})
        )

        duplicates_sheet = pd.concat(
            [
                validation_dm["duplicate_df"].assign(Source="D-Money"),
                validation_bank["duplicate_df"].assign(Source="Banque"),
            ],
            ignore_index=True,
        )

        kpis_export = {
            "Date du rapprochement": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Total transactions D-Money": stats["total_dm"],
            "Total transactions Banque": stats["total_bank"],
            "Transactions rapprochées": stats["matched"],
            "Uniquement D-Money": len(results["only_dm"]),
            "Uniquement Banque": len(results["only_bank"]),
            "Total écarts": stats["ecarts"],
            "Taux de rapprochement (%)": stats["taux"],
            "Doublons D-Money (lignes)": validation_dm["duplicate_row_count"],
            "Doublons Banque (lignes)": validation_bank["duplicate_row_count"],
        }
        if amounts_df is not None:
            kpis_export["Montant total D-Money"] = round(amounts_df["Amount_DM_num"].sum(), 2)
            kpis_export["Montant total Banque"] = round(amounts_df["Amount_BANK_num"].sum(), 2)
            kpis_export["Écart total"] = round(amounts_df["Amount_DM_num"].sum() - amounts_df["Amount_BANK_num"].sum(), 2)
            kpis_export["Montants incohérents"] = int((amounts_df["Statut_Montant"] == "Différent").sum())

        sheets = {
            "Rapprochées": results["matched"],
            "Uniquement D-Money": results["only_dm"],
            "Uniquement Banque": results["only_bank"],
            "Differences montants": diff_amounts_sheet,
            "Doublons": duplicates_sheet,
        }

        excel_buffer = export_to_excel(sheets, kpis_export)

        st.download_button(
            label="⬇️ Télécharger le rapport Excel complet",
            data=excel_buffer,
            file_name=f"rapprochement_dmoney_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
